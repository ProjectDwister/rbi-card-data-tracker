"""
RBI Card Data Historical Backfill
Target: Jan 2017 to Jul 2023

Outputs are saved in the same folder as this script:
- rbi_card_historical_backfill_2017_to_2023.xlsx
- rbi_card_month_index_historical.csv
- rbi_card_raw_historical.csv
- rbi_card_wide_historical.csv
- rbi_card_backfill_log.csv

Important:
Some old RBI rbidocs .XLS links close the connection when requested by Python.
For those months, the script supports a manual fallback folder:

    manual_sources/
      2017-01.xls
      2017-02.xls
      ...

If a month cannot be downloaded automatically, download the RBI XLS/PDF manually
from browser and place it in manual_sources using the month_key in the file name.
The script will pick it up automatically on the next run.
"""

import hashlib
import re
import time
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from urllib.parse import urljoin
from urllib.request import Request, urlopen

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


BASE_ARCHIVE_URL = "https://www.rbi.org.in/Scripts/ATMView.aspx"
BASE_DETAIL_URL = "https://www.rbi.org.in/Scripts/ATMView.aspx?atmid={atmid}"
RBI_BASE = "https://www.rbi.org.in"

OUTPUT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
MANUAL_SOURCES_DIR = OUTPUT_DIR / "manual_sources"
MANUAL_SOURCES_DIR.mkdir(parents=True, exist_ok=True)
DEBUG_DIR = OUTPUT_DIR / "debug_payloads"
DEBUG_DIR.mkdir(parents=True, exist_ok=True)

OUT_XLSX = OUTPUT_DIR / "rbi_card_historical_backfill_2017_to_2023.xlsx"
OUT_INDEX_CSV = OUTPUT_DIR / "rbi_card_month_index_historical.csv"
OUT_RAW_CSV = OUTPUT_DIR / "rbi_card_raw_historical.csv"
OUT_WIDE_CSV = OUTPUT_DIR / "rbi_card_wide_historical.csv"
OUT_LOG_CSV = OUTPUT_DIR / "rbi_card_backfill_log.csv"

START_YEAR = 2017
START_MONTH = 1
END_YEAR = 2023
END_MONTH = 7

FALLBACK_SCAN_MIN_ATMID = 1
FALLBACK_SCAN_MAX_ATMID = 230

SLEEP_SECONDS = 0.35
REQUEST_TIMEOUT = 45

MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}

MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
    7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December",
}

WIDE_COLUMNS = [
    "Sr No", "Bank Name",
    "ATMs & CRMs - On-site", "ATMs & CRMs - Off-site", "PoS", "Micro ATMs",
    "Bharat QR Codes", "UPI QR Codes", "Credit Cards Outstanding", "Debit Cards Outstanding",
    "Credit Card - Card Payments at PoS - Volume", "Credit Card - Card Payments at PoS - Value Rs000",
    "Credit Card - Online e-com - Volume", "Credit Card - Online e-com - Value Rs000",
    "Credit Card - Others - Volume", "Credit Card - Others - Value Rs000",
    "Credit Card - Cash Withdrawal at ATM - Volume", "Credit Card - Cash Withdrawal at ATM - Value Rs000",
    "Debit Card - Card Payments at PoS - Volume", "Debit Card - Card Payments at PoS - Value Rs000",
    "Debit Card - Online e-com - Volume", "Debit Card - Online e-com - Value Rs000",
    "Debit Card - Others - Volume", "Debit Card - Others - Value Rs000",
    "Debit Card - Cash Withdrawal ATM - Volume", "Debit Card - Cash Withdrawal ATM - Value Rs000",
    "Debit Card - Cash Withdrawal PoS - Volume", "Debit Card - Cash Withdrawal PoS - Value Rs000",
]


def make_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=3,
        connect=3,
        read=3,
        backoff_factor=0.8,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Connection": "close",
    })
    return session


def log(log_rows, level, func, message, details=""):
    details = str(details or "")
    if len(details) > 3000:
        details = details[:3000] + " ... [truncated]"
    print(f"{datetime.now():%Y-%m-%d %H:%M:%S}  {level:<5}  {func:<34}  {message}")
    log_rows.append({
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "level": level,
        "function": func,
        "message": message,
        "details": details,
    })


def make_month_key(year, month_no):
    return f"{int(year)}-{int(month_no):02d}"


def expected_month_keys():
    return [str(p) for p in pd.period_range(
        start=f"{START_YEAR}-{START_MONTH:02d}",
        end=f"{END_YEAR}-{END_MONTH:02d}",
        freq="M",
    )]


def in_target_range(year, month_no):
    ym = int(year) * 100 + int(month_no)
    return START_YEAR * 100 + START_MONTH <= ym <= END_YEAR * 100 + END_MONTH


def clean_cell(x):
    if x is None:
        return ""
    s = str(x).replace("\xa0", " ").replace("\u200b", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if re.fullmatch(r"-?\d+\.0", s):
        s = s[:-2]
    return s


def clean_text_from_html(html):
    return BeautifulSoup(html or "", "html.parser").get_text(" ", strip=True)


def fetch_text(session, url):
    response = session.get(
        url,
        headers={
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": BASE_ARCHIVE_URL,
        },
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    response.encoding = response.apparent_encoding or "utf-8"
    return response.text


def fetch_bytes_with_requests(session, url):
    response = session.get(
        url,
        headers={
            "Accept": "application/vnd.ms-excel,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/html,*/*",
            "Referer": BASE_ARCHIVE_URL,
        },
        timeout=REQUEST_TIMEOUT,
        allow_redirects=True,
    )
    response.raise_for_status()
    return response.content, response.url, response.headers.get("content-type", "")


def fetch_bytes_with_urllib(url):
    req = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
            "Accept": "application/vnd.ms-excel,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/html,*/*",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": BASE_ARCHIVE_URL,
            "Connection": "close",
        },
    )
    with urlopen(req, timeout=REQUEST_TIMEOUT) as response:
        return response.read(), response.geturl(), response.headers.get("content-type", "")


def fetch_bytes(session, url):
    try:
        return fetch_bytes_with_requests(session, url)[0]
    except Exception as first_error:
        try:
            return fetch_bytes_with_urllib(url)[0]
        except Exception as second_error:
            raise RuntimeError(f"requests_error={repr(first_error)}, urllib_error={repr(second_error)}")


def parse_month_year_from_text(text):
    text = re.sub(r"\s+", " ", str(text or " ")).strip()
    patterns = [
        r"for\s+the\s+Month\s+of\s+([A-Za-z]+)\s+(\d{4})",
        r"for\s+the\s+Month\s+of\s+([A-Za-z]+)[\s\-']+(\d{2})",
        r"Bank[- ]wise\s+ATM/POS/Card\s+Statistics\s*[-–]\s*([A-Za-z]+)\s+(\d{4})",
        r"ATM\s*,?\s*Acceptance\s+Infrastructure\s+and\s+Card\s+Statistics\s+for\s+the\s+Month\s+of\s+([A-Za-z]+)\s+(\d{4})",
        r"ATM\s*&\s*Card\s+Statistics\s+for\s+([A-Za-z]+)\s*[-–]\s*(\d{4})",
        r"\b([A-Za-z]+)\s*[-–]\s*(\d{4})\b",
        r"\b([A-Za-z]+)\s*[-–]\s*(\d{2})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        month_no = MONTHS.get(match.group(1).lower())
        if not month_no:
            continue
        year = int(match.group(2))
        if year < 100:
            year += 2000
        return year, month_no, MONTH_NAMES[month_no]
    return None


def parse_month_year_from_html(html):
    soup = BeautifulSoup(html or "", "html.parser")
    for tag in soup.find_all(["span", "h1", "h2", "h3", "td", "th", "caption"]):
        txt = tag.get_text(" ", strip=True)
        if "card statistics" in txt.lower() or "month of" in txt.lower():
            parsed = parse_month_year_from_text(txt)
            if parsed:
                return parsed
    return parse_month_year_from_text(clean_text_from_html(html))


def is_revised_text(text):
    return bool(re.search(r"\brevised\b", str(text or ""), flags=re.IGNORECASE))


def extract_hidden_form_fields(html):
    soup = BeautifulSoup(html or "", "html.parser")
    data = {}
    for inp in soup.find_all("input"):
        name = inp.get("name")
        if name:
            data[name] = inp.get("value", "")
    return data


def post_archive_for_month(session, base_html, year, month_no):
    data = extract_hidden_form_fields(base_html)
    data["hdnYear"] = str(year)
    data["hdnMonth"] = str(month_no)
    response = session.post(
        BASE_ARCHIVE_URL,
        data=data,
        headers={
            "Referer": BASE_ARCHIVE_URL,
            "Origin": "https://www.rbi.org.in",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    response.encoding = response.apparent_encoding or "utf-8"
    return response.text


def extract_links_from_archive_html(html, target_year=None, target_month_no=None):
    soup = BeautifulSoup(html or "", "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = a.get("href", "")
        text = " ".join([a.get_text(" ", strip=True), a.get("title", ""), a.get("aria-label", ""), href])
        full_url = urljoin(RBI_BASE, href).replace("&amp;", "&")
        is_detail = bool(re.search(r"ATMView\.aspx\?atmid=\d+", full_url, flags=re.IGNORECASE))
        is_excel = bool(re.search(r"\.xlsx?($|\?)", full_url, flags=re.IGNORECASE))
        is_pdf = bool(re.search(r"\.pdf($|\?)", full_url, flags=re.IGNORECASE))
        if not (is_detail or is_excel or is_pdf):
            continue
        parsed = parse_month_year_from_text(text)
        month_key = ""
        if parsed:
            y, m, _ = parsed
            month_key = make_month_key(y, m)
        if target_year and target_month_no:
            target_key = make_month_key(target_year, target_month_no)
            if month_key and month_key != target_key:
                continue
            if not month_key:
                month_key = target_key
        links.append({
            "url": full_url,
            "text": text,
            "is_detail": is_detail,
            "is_excel": is_excel,
            "is_pdf": is_pdf,
            "month_key": month_key,
        })
    return links


def add_link_to_discovery(record, link):
    url = link["url"]
    if link["is_detail"]:
        record["page_url"] = url
        m = re.search(r"atmid=(\d+)", url, flags=re.IGNORECASE)
        if m:
            record["atmid"] = int(m.group(1))
    if link["is_excel"]:
        record["excel_url"] = url
    if link["is_pdf"]:
        record["pdf_url"] = url


def discover_month_pages_from_archive(log_rows):
    session = make_session()
    discovered = {}
    log(log_rows, "INFO", "discover_archive", "Fetching RBI archive page")
    base_html = fetch_text(session, BASE_ARCHIVE_URL)
    for key in expected_month_keys():
        year = int(key[:4])
        month_no = int(key[5:7])
        try:
            html = post_archive_for_month(session, base_html, year, month_no)
            links = extract_links_from_archive_html(html, year, month_no)
            rec = {
                "month_key": key,
                "year": year,
                "month_no": month_no,
                "month_name": MONTH_NAMES[month_no],
                "source": "archive_post_html",
                "archive_html": html,
                "atmid": "",
                "page_url": "",
                "excel_url": "",
                "pdf_url": "",
            }
            for link in links:
                add_link_to_discovery(rec, link)
            discovered[key] = rec
            log(log_rows, "INFO", "discover_archive", f"Discovered archive HTML for {key}", f"links={len(links)}, html_length={len(html)}")
        except Exception as e:
            log(log_rows, "ERROR", "discover_archive", f"Archive POST failed for {key}", repr(e))
        time.sleep(SLEEP_SECONDS)
    return discovered


def extract_first_excel_from_html(html):
    soup = BeautifulSoup(html or "", "html.parser")
    for a in soup.find_all("a", href=True):
        href = a.get("href", "")
        if re.search(r"\.xlsx?($|\?)", href, flags=re.IGNORECASE):
            return urljoin(RBI_BASE, href).replace("&amp;", "&")
    return ""


def discover_from_direct_atmid_scan(log_rows):
    session = make_session()
    found = {}
    log(log_rows, "INFO", "discover_atmid", f"Fallback scanning atmid {FALLBACK_SCAN_MIN_ATMID} to {FALLBACK_SCAN_MAX_ATMID}")
    for atmid in range(FALLBACK_SCAN_MIN_ATMID, FALLBACK_SCAN_MAX_ATMID + 1):
        url = BASE_DETAIL_URL.format(atmid=atmid)
        try:
            html = fetch_text(session, url)
            parsed = parse_month_year_from_html(html)
            if not parsed:
                continue
            year, month_no, month_name = parsed
            if not in_target_range(year, month_no):
                continue
            month_key = make_month_key(year, month_no)
            if month_key not in found:
                found[month_key] = {
                    "month_key": month_key,
                    "year": year,
                    "month_no": month_no,
                    "month_name": month_name,
                    "atmid": atmid,
                    "page_url": url,
                    "excel_url": extract_first_excel_from_html(html),
                    "pdf_url": "",
                    "source": "direct_atmid_scan",
                }
                log(log_rows, "INFO", "discover_atmid", f"Found {month_key} at atmid={atmid}")
        except Exception as e:
            log(log_rows, "WARN", "discover_atmid", f"Failed atmid={atmid}", repr(e))
        time.sleep(0.15)
    return found


def parse_html_table_rows(table):
    rows = []
    for tr in table.find_all("tr"):
        cells = [clean_cell(cell.get_text(" ", strip=True)) for cell in tr.find_all(["td", "th"])]
        if "".join(cells).strip():
            rows.append(cells)
    return normalize_rows(rows)


def normalize_rows(rows):
    if not rows:
        return []
    max_cols = max(len(r) for r in rows)
    out = []
    for row in rows:
        r = list(row)
        while len(r) < max_cols:
            r.append("")
        out.append(r)
    return out


def choose_best_html_table(html):
    soup = BeautifulSoup(html or "", "html.parser")
    tables = soup.find_all("table")
    best_rows = []
    best_score = -1
    keywords = [
        "Scheduled Commercial Banks", "Public Sector Banks", "Private Sector Banks", "Foreign Banks",
        "Credit Card", "Debit Card", "ATMs", "PoS", "ATM, Acceptance Infrastructure", "Card Payments",
    ]
    for table in tables:
        text = table.get_text(" ", strip=True)
        rows = parse_html_table_rows(table)
        if not rows:
            continue
        max_cols = max(len(r) for r in rows)
        if len(rows) < 20 or max_cols < 10:
            continue
        keyword_hits = sum(1 for k in keywords if k.lower() in text.lower())
        if keyword_hits < 2:
            continue
        bank_rows = len(extract_bank_rows(rows, quiet=True))
        score = bank_rows * 10000 + keyword_hits * 1000 + len(rows) * max_cols
        if score > best_score:
            best_score = score
            best_rows = rows
    return best_rows


def trim_leading_blank_cells(vals):
    vals = list(vals)
    while vals and str(vals[0]).strip() == "":
        vals.pop(0)
    return vals


def is_integer_like(value):
    return bool(re.fullmatch(r"\d+", str(value).strip()))


def is_numberish(value):
    s = clean_cell(value).replace(",", "")
    return bool(s and re.fullmatch(r"-?\d+(\.\d+)?", s))


def to_number_or_blank(value):
    s = clean_cell(value).replace(",", "")
    if s == "":
        return ""
    try:
        n = float(s)
        return int(n) if n.is_integer() else n
    except Exception:
        return clean_cell(value)


def lakh_to_rs000(value):
    s = clean_cell(value).replace(",", "")
    if s == "":
        return ""
    try:
        n = float(s) * 100
        return int(n) if n.is_integer() else n
    except Exception:
        return clean_cell(value)


def is_bank_group_row(row_values):
    vals = trim_leading_blank_cells(row_values)
    if not vals:
        return False
    if is_integer_like(vals[0]):
        return False
    joined = " ".join(str(x) for x in vals).strip()
    patterns = [
        "Scheduled Commercial Banks", "Public Sector Banks", "Private Sector Banks", "Foreign Banks",
        "Payment Banks", "Payments Banks", "Small Finance Banks", "Regional Rural Banks",
        "Co-operative Banks", "Cooperative Banks",
    ]
    return any(p.lower() in joined.lower() for p in patterns)


def looks_like_old_format_bank_row(vals):
    vals = trim_leading_blank_cells(vals)
    if len(vals) < 10:
        return False
    bank_name = clean_cell(vals[0])
    if not bank_name:
        return False
    bad_names = [
        "bank name", "scheduled commercial banks", "public sector banks", "private sector banks",
        "foreign banks", "payment banks", "payments banks", "small finance banks", "total", "note",
        "atm & card statistics", "number of atm", "no. of outstanding",
    ]
    if any(x in bank_name.lower() for x in bad_names):
        return False
    numeric_count = sum(1 for x in vals[1:15] if is_numberish(x))
    return numeric_count >= 8


def extract_old_format_bank_row(vals, bank_group, generated_sr_no):
    vals = trim_leading_blank_cells(vals)
    while len(vals) < 15:
        vals.append("")
    pos_online = to_number_or_blank(vals[3])
    pos_offline = to_number_or_blank(vals[4])
    try:
        total_pos = float(pos_online or 0) + float(pos_offline or 0)
        total_pos = int(total_pos) if total_pos.is_integer() else total_pos
    except Exception:
        total_pos = ""
    record = {col: "" for col in WIDE_COLUMNS}
    record["Sr No"] = generated_sr_no
    record["Bank Name"] = clean_cell(vals[0])
    record["ATMs & CRMs - On-site"] = to_number_or_blank(vals[1])
    record["ATMs & CRMs - Off-site"] = to_number_or_blank(vals[2])
    record["PoS"] = total_pos
    record["Credit Cards Outstanding"] = to_number_or_blank(vals[5])
    record["Debit Cards Outstanding"] = to_number_or_blank(vals[10])
    record["Credit Card - Cash Withdrawal at ATM - Volume"] = to_number_or_blank(vals[6])
    record["Credit Card - Cash Withdrawal at ATM - Value Rs000"] = lakh_to_rs000(vals[8])
    record["Credit Card - Card Payments at PoS - Volume"] = to_number_or_blank(vals[7])
    record["Credit Card - Card Payments at PoS - Value Rs000"] = lakh_to_rs000(vals[9])
    record["Debit Card - Cash Withdrawal ATM - Volume"] = to_number_or_blank(vals[11])
    record["Debit Card - Cash Withdrawal ATM - Value Rs000"] = lakh_to_rs000(vals[13])
    record["Debit Card - Card Payments at PoS - Volume"] = to_number_or_blank(vals[12])
    record["Debit Card - Card Payments at PoS - Value Rs000"] = lakh_to_rs000(vals[14])
    record["bank_group"] = bank_group
    return record


def extract_bank_rows(rows, quiet=False):
    wide_rows = []
    bank_group = ""
    generated_sr_no = 1
    for row in rows:
        vals = [clean_cell(v) for v in row]
        vals = trim_leading_blank_cells(vals)
        if not vals:
            continue
        if is_bank_group_row(vals):
            bank_group = " ".join(vals).strip()
            continue
        if len(vals) >= 2 and is_integer_like(vals[0]) and vals[1]:
            data = vals[:len(WIDE_COLUMNS)]
            while len(data) < len(WIDE_COLUMNS):
                data.append("")
            record = {"bank_group": bank_group}
            for col_name, value in zip(WIDE_COLUMNS, data):
                record[col_name] = clean_cell(value)
            wide_rows.append(record)
            continue
        if looks_like_old_format_bank_row(vals):
            wide_rows.append(extract_old_format_bank_row(vals, bank_group, generated_sr_no))
            generated_sr_no += 1
    return wide_rows


def rows_from_dataframe(df):
    df = df.fillna("")
    rows = []
    for _, row in df.iterrows():
        vals = [clean_cell(v) for v in row.tolist()]
        while vals and vals[-1] == "":
            vals.pop()
        if "".join(vals).strip():
            rows.append(vals)
    return normalize_rows(rows)


def parse_html_bytes_table_rows(content):
    encodings = ["utf-8-sig", "utf-8", "utf-16", "utf-16le", "cp1252", "latin1"]
    last_error = None
    for enc in encodings:
        try:
            html = content.decode(enc, errors="ignore")
            if "<table" not in html.lower():
                continue
            rows = choose_best_html_table(html)
            if rows:
                return rows
            try:
                dfs = pd.read_html(StringIO(html))
                best_rows = []
                best_score = -1
                for df in dfs:
                    rows2 = rows_from_dataframe(df)
                    if not rows2:
                        continue
                    bank_rows = extract_bank_rows(rows2, quiet=True)
                    max_cols = max(len(r) for r in rows2)
                    score = len(bank_rows) * 10000 + len(rows2) * max_cols
                    if score > best_score:
                        best_score = score
                        best_rows = rows2
                if best_rows:
                    return best_rows
            except Exception as e:
                last_error = e
        except Exception as e:
            last_error = e
    raise RuntimeError(f"Could not parse HTML/pseudo-Excel content. Last error={repr(last_error)}")


def parse_excel_table_rows_from_bytes(content):
    first_part = content[:2000].lower()
    if first_part.strip().startswith(b"<!doctype") or b"<html" in first_part or b"<table" in first_part:
        return parse_html_bytes_table_rows(content)
    if content[:2] == b"PK":
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            vals = [clean_cell(v) for v in row]
            while vals and vals[-1] == "":
                vals.pop()
            if "".join(vals).strip():
                rows.append(vals)
        return normalize_rows(rows)
    try:
        df = pd.read_excel(BytesIO(content), header=None, dtype=str, engine="xlrd")
        return rows_from_dataframe(df)
    except Exception as xls_err:
        try:
            return parse_html_bytes_table_rows(content)
        except Exception as html_err:
            raise RuntimeError(f"Could not parse as XLS/XLSX/HTML. xls_err={repr(xls_err)}, html_err={repr(html_err)}")


def parse_excel_table_rows(session, excel_url):
    content = fetch_bytes(session, excel_url)
    return parse_excel_table_rows_from_bytes(content)


def find_manual_source_file(month_key):
    patterns = [
        f"{month_key}.*",
        f"{month_key.replace('-', '_')}.*",
        f"{month_key.replace('-', '')}.*",
    ]
    for pattern in patterns:
        for path in sorted(MANUAL_SOURCES_DIR.glob(pattern)):
            if path.suffix.lower() in [".xls", ".xlsx", ".html", ".htm", ".txt", ".csv"]:
                return path
    return None


def parse_manual_source_file(path):
    suffix = path.suffix.lower()
    content = path.read_bytes()
    if suffix == ".csv":
        df = pd.read_csv(path, header=None, dtype=str, encoding="utf-8-sig")
        return rows_from_dataframe(df)
    return parse_excel_table_rows_from_bytes(content)


def build_raw_records(rows, meta):
    records = []
    max_cols = max([len(r) for r in rows], default=1)
    for idx, row in enumerate(rows):
        rec = {**meta, "source_row_no": idx + 1}
        for i in range(max_cols):
            rec[f"raw_col_{i + 1}"] = clean_cell(row[i]) if i < len(row) else ""
        records.append(rec)
    return records


def merge_discoveries(archive_discovered, atmid_discovered):
    merged = {}
    for key in expected_month_keys():
        archive_rec = archive_discovered.get(key, {})
        atmid_rec = atmid_discovered.get(key, {})
        if not archive_rec and not atmid_rec:
            continue
        rec = {}
        rec.update(archive_rec)
        for field in ["atmid", "page_url", "excel_url", "pdf_url", "source"]:
            if not rec.get(field) and atmid_rec.get(field):
                rec[field] = atmid_rec.get(field)
        if archive_rec.get("archive_html"):
            rec["archive_html"] = archive_rec.get("archive_html")
        if atmid_rec.get("page_url"):
            rec["page_url"] = atmid_rec.get("page_url")
            rec["atmid"] = atmid_rec.get("atmid", rec.get("atmid", ""))
            rec["source"] = "archive_post_html_plus_direct_atmid"
        year = int(key[:4])
        month_no = int(key[5:7])
        rec.setdefault("month_key", key)
        rec.setdefault("year", year)
        rec.setdefault("month_no", month_no)
        rec.setdefault("month_name", MONTH_NAMES[month_no])
        rec.setdefault("atmid", "")
        rec.setdefault("page_url", "")
        rec.setdefault("excel_url", "")
        rec.setdefault("pdf_url", "")
        rec.setdefault("source", "")
        merged[key] = rec
    return merged


def parse_one_month(session, key, rec, log_rows):
    archive_html = rec.get("archive_html", "")
    page_url = rec.get("page_url", "")
    excel_url = rec.get("excel_url", "")

    manual_file = find_manual_source_file(key)
    if manual_file:
        rows = parse_manual_source_file(manual_file)
        if rows:
            return rows, f"manual_source:{manual_file.name}", excel_url

    if page_url:
        try:
            html = fetch_text(session, page_url)
            rows = choose_best_html_table(html)
            if rows:
                if not excel_url:
                    excel_url = extract_first_excel_from_html(html)
                return rows, "detail_html", excel_url
            if not excel_url:
                excel_url = extract_first_excel_from_html(html)
        except Exception as page_err:
            log(log_rows, "WARN", "import", f"Detail page parse failed for {key}", repr(page_err))

    if excel_url:
        try:
            rows = parse_excel_table_rows(session, excel_url)
            if rows:
                return rows, "excel_or_excel_html", excel_url
        except Exception as excel_err:
            log(log_rows, "WARN", "import", f"Excel fallback failed for {key}", f"url={excel_url}, error={repr(excel_err)}")

    if archive_html:
        rows = choose_best_html_table(archive_html)
        if rows:
            return rows, "archive_html", excel_url

    raise RuntimeError("No real data table parsed from detail HTML, Excel URL, manual file, or archive HTML")


def import_discovered_months(discovered, log_rows):
    session = make_session()
    month_index_rows = []
    raw_records = []
    wide_records = []

    for key in expected_month_keys():
        rec = discovered.get(key)
        if not rec:
            log(log_rows, "ERROR", "import", f"No discovered source for {key}")
            continue
        year = int(rec["year"])
        month_no = int(rec["month_no"])
        month_name = rec["month_name"]
        month_date = f"{year}-{month_no:02d}-01"
        fetched_at = datetime.now().isoformat(timespec="seconds")

        try:
            rows, source_used, excel_url = parse_one_month(session, key, rec, log_rows)
            wide_rows = extract_bank_rows(rows)
            if len(wide_rows) < 10:
                sample = rows[:8]
                raise RuntimeError(f"Too few bank rows. raw_rows={len(rows)}, bank_rows={len(wide_rows)}, source={source_used}, sample={sample}")

            meta = {
                "month_key": key,
                "month_date": month_date,
                "year": year,
                "month_no": month_no,
                "month_name": month_name,
                "atmid": rec.get("atmid", ""),
                "is_revised": is_revised_text(" ".join([str(rec.get("text", "")), excel_url, rec.get("page_url", ""), rec.get("archive_html", "")[:1000]])),
                "page_url": rec.get("page_url", ""),
                "excel_url": excel_url,
                "fetched_at": fetched_at,
                "source_used": source_used,
            }
            raw_records.extend(build_raw_records(rows, meta))
            for br in wide_rows:
                wide_records.append({
                    "month_key": key,
                    "month_date": month_date,
                    "year": year,
                    "month_no": month_no,
                    "month_name": month_name,
                    "atmid": rec.get("atmid", ""),
                    "is_revised": meta["is_revised"],
                    "page_url": rec.get("page_url", ""),
                    "excel_url": excel_url,
                    "fetched_at": fetched_at,
                    "source_used": source_used,
                    **br,
                })
            page_hash_text = "|".join([key, rec.get("page_url", ""), excel_url, str(len(rows)), str(len(wide_rows)), source_used])
            month_index_rows.append({
                "month_key": key,
                "month_date": month_date,
                "year": year,
                "month_no": month_no,
                "month_name": month_name,
                "atmid": rec.get("atmid", ""),
                "is_revised": meta["is_revised"],
                "page_url": rec.get("page_url", ""),
                "excel_url": excel_url,
                "pdf_url": rec.get("pdf_url", ""),
                "source": rec.get("source", ""),
                "source_used": source_used,
                "page_hash": hashlib.md5(page_hash_text.encode("utf-8")).hexdigest(),
                "status": "IMPORTED",
                "fetched_at": fetched_at,
                "raw_rows": len(rows),
                "bank_rows": len(wide_rows),
            })
            log(log_rows, "INFO", "import", f"Imported {key}", f"source={source_used}, raw_rows={len(rows)}, bank_rows={len(wide_rows)}")
        except Exception as e:
            log(log_rows, "ERROR", "import", f"Failed to import {key}", repr(e))
        time.sleep(SLEEP_SECONDS)

    month_index_df = pd.DataFrame(month_index_rows)
    raw_df = pd.DataFrame(raw_records)
    wide_df = pd.DataFrame(wide_records)
    if not month_index_df.empty:
        month_index_df = month_index_df.sort_values(["year", "month_no"]).reset_index(drop=True)
    if not raw_df.empty:
        raw_df = raw_df.sort_values(["year", "month_no", "source_row_no"]).reset_index(drop=True)
    if not wide_df.empty:
        wide_df = wide_df.sort_values(["year", "month_no", "bank_group", "Sr No"]).reset_index(drop=True)
    return month_index_df, raw_df, wide_df


def validate_outputs(month_index_df, raw_df, wide_df):
    expected_keys = expected_month_keys()
    found_keys = set(month_index_df["month_key"].astype(str)) if not month_index_df.empty else set()
    missing = [m for m in expected_keys if m not in found_keys]
    extra = [m for m in sorted(found_keys) if m not in expected_keys]
    summary_rows = []
    for key in expected_keys:
        raw_count = int((raw_df["month_key"].astype(str) == key).sum()) if not raw_df.empty else 0
        wide_count = int((wide_df["month_key"].astype(str) == key).sum()) if not wide_df.empty else 0
        summary_rows.append({
            "month_key": key,
            "raw_rows": raw_count,
            "bank_rows": wide_count,
            "status": "OK" if key in found_keys else "MISSING",
        })
    validation_df = pd.DataFrame(summary_rows)
    print("\n================ VALIDATION ================")
    print(f"Expected months : {len(expected_keys)}")
    print(f"Found months    : {len(found_keys)}")
    print(f"Missing months  : {len(missing)}")
    print(f"Extra months    : {len(extra)}")
    if missing:
        print("Missing:", ", ".join(missing))
    if extra:
        print("Extra:", ", ".join(extra))
    print("============================================\n")
    return validation_df


def safe_to_excel(df, writer, sheet_name):
    if df.empty:
        pd.DataFrame({"notice": ["No rows"]}).to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def save_outputs(month_index_df, raw_df, wide_df, log_df, validation_df):
    month_index_df.to_csv(OUT_INDEX_CSV, index=False, encoding="utf-8-sig")
    raw_df.to_csv(OUT_RAW_CSV, index=False, encoding="utf-8-sig")
    wide_df.to_csv(OUT_WIDE_CSV, index=False, encoding="utf-8-sig")
    log_df.to_csv(OUT_LOG_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        safe_to_excel(month_index_df, writer, "Month_Index_Historical")
        safe_to_excel(raw_df, writer, "Raw_Historical")
        safe_to_excel(wide_df, writer, "Wide_Historical")
        safe_to_excel(validation_df, writer, "Validation")
        safe_to_excel(log_df, writer, "Backfill_Log")
        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for col_idx in range(1, min(ws.max_column, 25) + 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                max_len = 12
                for row_idx in range(1, min(ws.max_row, 200) + 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, min(len(str(val)), 50))
                ws.column_dimensions[col_letter].width = max_len + 2
    print(f"Saved Excel : {OUT_XLSX}")
    print(f"Saved CSVs  : {OUTPUT_DIR}")


def debug_download_payloads():
    session = make_session()
    test_urls = {
        "2017-01_failed_xls": "https://rbidocs.rbi.org.in/rdocs/ATM/DOCs/ATMCS012017849801C2A75D447F9C7E683BDC7B1A8D.XLS",
        "2020-02_suspicious_detail": "https://www.rbi.org.in/Scripts/ATMView.aspx?atmid=108",
        "2023-07_good_detail": "https://www.rbi.org.in/Scripts/ATMView.aspx?atmid=149",
    }
    for name, url in test_urls.items():
        print("\n========================================")
        print("Testing:", name)
        print("URL:", url)
        try:
            content = fetch_bytes(session, url)
            print("Content-Length:", len(content))
            print("First 100 bytes:", content[:100])
            raw_path = DEBUG_DIR / f"{name}.bin"
            raw_path.write_bytes(content)
            print("Saved raw payload:", raw_path)
            try:
                rows = parse_excel_table_rows_from_bytes(content)
                wide_rows = extract_bank_rows(rows)
                print("parsed rows:", len(rows))
                print("bank rows:", len(wide_rows))
                for r in rows[:5]:
                    print(r[:10])
            except Exception as e:
                print("Parser error:", repr(e))
        except Exception as e:
            print("Download error:", repr(e))


def main():
    print("====================================================")
    print("RBI Card Data Historical Backfill")
    print(f"Target range: {START_YEAR}-{START_MONTH:02d} to {END_YEAR}-{END_MONTH:02d}")
    print("Method: archive discovery + direct detail pages + Excel/HTML + manual_sources fallback")
    print("====================================================")
    log_rows = []
    archive_discovered = discover_month_pages_from_archive(log_rows)
    atmid_discovered = discover_from_direct_atmid_scan(log_rows)
    discovered = merge_discoveries(archive_discovered, atmid_discovered)
    log(log_rows, "INFO", "main", "Discovery complete", f"archive={len(archive_discovered)}, atmid={len(atmid_discovered)}, merged={len(discovered)}")
    month_index_df, raw_df, wide_df = import_discovered_months(discovered, log_rows)
    validation_df = validate_outputs(month_index_df, raw_df, wide_df)
    log_df = pd.DataFrame(log_rows)
    save_outputs(month_index_df, raw_df, wide_df, log_df, validation_df)


if __name__ == "__main__":
    main()
